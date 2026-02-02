from flask import render_template, request, redirect, url_for, session, flash, send_file
import matplotlib
matplotlib.use('Agg')  # Use non-interactive backend for Flask
import matplotlib.pyplot as plt
import os
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

def admin(mysql):
    if 'username' not in session:
        return redirect(url_for('login'))
    
    if session.get('admin') != 1:
        flash("You do not have admin privileges")
        return redirect(url_for('home'))
    
    # Get statistics for dashboard
    cur = mysql.connection.cursor()
    
    # Get total customers
    cur.execute("SELECT COUNT(*) FROM users WHERE admin = 0")
    total_customers = cur.fetchone()[0]
    
    # Get total products
    cur.execute("SELECT COUNT(*) FROM produk")
    total_products = cur.fetchone()[0]
    
    # Get total orders
    cur.execute("SELECT COUNT(*) FROM history")
    total_orders = cur.fetchone()[0]
    
    # Get total revenue
    cur.execute("SELECT SUM(jumlah_pembayaran) FROM history WHERE status_produk != 'Cancelled'")
    total_revenue = cur.fetchone()[0] or 0
    
    # Get merk data for chart
    cur.execute("SELECT nama_merk, jumlahpenjualan, keuntungan FROM merk ORDER BY jumlahpenjualan DESC")
    merks_data = cur.fetchall()
    cur.close()
    
    # Generate charts if there's merk data with sales > 0
    has_merks = False
    if merks_data:
        nama_merk_list = [m[0] for m in merks_data]
        jumlahpenjualan_list = [m[1] for m in merks_data]
        keuntungan_list = [m[2] for m in merks_data]
        
        # Check if there's any data > 0
        if any(jumlahpenjualan > 0 for jumlahpenjualan in jumlahpenjualan_list):
            has_merks = True
            static_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'static')
            if not os.path.exists(static_dir):
                os.makedirs(static_dir)
            
            # Bar chart for sales
            plt.figure(figsize=(10, 6))
            plt.bar(nama_merk_list, jumlahpenjualan_list, color='steelblue')
            plt.title('Jumlah Penjualan per Merk')
            plt.xlabel('Nama Merk')
            plt.ylabel('Jumlah Penjualan')
            plt.xticks(rotation=45, ha='right')
            plt.tight_layout()
            plt.savefig(os.path.join(static_dir, 'merk_bar.png'), dpi=100, bbox_inches='tight')
            plt.close()
            
            # Pie chart for sales proportion (only if there's data)
            plt.figure(figsize=(10, 8))
            plt.pie(jumlahpenjualan_list, labels=nama_merk_list, autopct='%1.1f%%', startangle=90)
            plt.title('Proporsi Penjualan per Merk')
            plt.tight_layout()
            plt.savefig(os.path.join(static_dir, 'merk_pie.png'), dpi=100, bbox_inches='tight')
            plt.close()
            
            # Bar chart for profit
            plt.figure(figsize=(10, 6))
            plt.bar(nama_merk_list, keuntungan_list, color='seagreen')
            plt.title('Keuntungan per Merk')
            plt.xlabel('Nama Merk')
            plt.ylabel('Keuntungan (Million)')
            plt.xticks(rotation=45, ha='right')
            plt.tight_layout()
            plt.savefig(os.path.join(static_dir, 'merk_profit.png'), dpi=100, bbox_inches='tight')
            plt.close()
    
    return render_template('admin.html', 
                         page_title='Dashboard',
                         active_page='dashboard',
                         username=session['username'],
                         total_customers=total_customers,
                         total_products=total_products,
                         total_orders=total_orders,
                         total_revenue=total_revenue,
                         has_merks=has_merks)

def merk_list(mysql):
    if 'username' not in session:
        return redirect(url_for('login'))
    
    if session.get('admin') != 1:
        flash("You do not have admin privileges")
        return redirect(url_for('home'))
    
    action = request.args.get('action', None)
    merk_id = request.args.get('merk_id', None)
    merk = None
    show_form = False
    

    if action == 'create':
        show_form = True
    elif action == 'edit' and merk_id:
        show_form = True

        cur = mysql.connection.cursor()
        cur.execute("SELECT merk_id, nama_merk, jumlahpenjualan, keuntungan FROM merk WHERE merk_id = %s", (merk_id,))
        row = cur.fetchone()
        cur.close()
        
        if row:
            merk = {
                'merk_id': row[0],
                'nama_merk': row[1],
                'jumlahpenjualan': row[2],
                'keuntungan': row[3]
            }
        else:
            flash("Merk not found", "error")
    

    if request.method == 'POST':
        nama_merk = request.form.get('nama_merk', '').strip()
        jumlahpenjualan = request.form.get('jumlahpenjualan', 0)
        keuntungan = request.form.get('keuntungan', 0)
        
        if not nama_merk:
            flash("Brand name is required", "error")
            return redirect(url_for('merk_list', action=action, merk_id=merk_id))
        
        try:
            cur = mysql.connection.cursor()
            if merk_id:

                cur.execute("UPDATE merk SET nama_merk = %s, jumlahpenjualan = %s, keuntungan = %s WHERE merk_id = %s",
                            (nama_merk, jumlahpenjualan, keuntungan, merk_id))
                flash("Merk updated successfully", "success")
            else:

                cur.execute("INSERT INTO merk (nama_merk, jumlahpenjualan, keuntungan) VALUES (%s, %s, %s)",
                            (nama_merk, jumlahpenjualan, keuntungan))
                flash("Merk created successfully", "success")
            
            mysql.connection.commit()
            cur.close()
            
            return redirect(url_for('merk_list'))
        except Exception as e:
            flash(f"Error: {str(e)}", "error")
            return redirect(url_for('merk_list', action=action, merk_id=merk_id))
    

    cur = mysql.connection.cursor()
    cur.execute("SELECT merk_id, nama_merk, jumlahpenjualan, keuntungan FROM merk ORDER BY merk_id DESC")
    merks_data = cur.fetchall()
    cur.close()
    
    merks = []
    for row in merks_data:
        merks.append({
            'merk_id': row[0],
            'nama_merk': row[1],
            'jumlahpenjualan': row[2],
            'keuntungan': row[3]
        })
    

    if merks:
        nama_merk_list = [m['nama_merk'] for m in merks]
        jumlahpenjualan_list = [m['jumlahpenjualan'] for m in merks]
        keuntungan_list = [m['keuntungan'] for m in merks]
        
        # Only generate charts if there's valid data (any value > 0)
        if any(jumlahpenjualan > 0 for jumlahpenjualan in jumlahpenjualan_list):
            static_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'static')
            if not os.path.exists(static_dir):
                os.makedirs(static_dir)
            
            # Bar chart for sales
            plt.figure(figsize=(10, 6))
            plt.bar(nama_merk_list, jumlahpenjualan_list, color='steelblue')
            plt.title('Jumlah Penjualan per Merk')
            plt.xlabel('Nama Merk')
            plt.ylabel('Jumlah Penjualan')
            plt.xticks(rotation=45, ha='right')
            plt.tight_layout()
            plt.savefig(os.path.join(static_dir, 'merk_bar.png'), dpi=100, bbox_inches='tight')
            plt.close()
            
            # Pie chart for sales proportion
            plt.figure(figsize=(10, 8))
            plt.pie(jumlahpenjualan_list, labels=nama_merk_list, autopct='%1.1f%%', startangle=90)
            plt.title('Proporsi Penjualan per Merk')
            plt.tight_layout()
            plt.savefig(os.path.join(static_dir, 'merk_pie.png'), dpi=100, bbox_inches='tight')
            plt.close()
    
    return render_template('merk.html', 
                         merks=merks, 
                         merk=merk,
                         show_form=show_form,
                         username=session['username'],
                         page_title='Sales Management - Merk',
                         active_page='sales')

def merk_delete(mysql, merk_id):
    if 'username' not in session:
        return redirect(url_for('login'))
    
    if session.get('admin') != 1:
        flash("You do not have admin privileges")
        return redirect(url_for('home'))
    
    try:
        cur = mysql.connection.cursor()
        cur.execute("DELETE FROM merk WHERE merk_id = %s", (merk_id,))
        mysql.connection.commit()
        cur.close()
        
        flash("Merk deleted successfully", "success")
    except Exception as e:
        flash(f"Error: {str(e)}", "error")
    
    return redirect(url_for('merk_list'))

def customer_list(mysql):
    if 'username' not in session:
        return redirect(url_for('login'))
    
    if session.get('admin') != 1:
        flash("You do not have admin privileges")
        return redirect(url_for('home'))
    

    cur = mysql.connection.cursor()
    cur.execute("SELECT id, username, first_name, last_name, email, phone, address FROM users WHERE admin = 0 ORDER BY id DESC")
    customers_data = cur.fetchall()
    cur.close()
    
    customers = []
    for row in customers_data:
        customers.append({
            'id': row[0],
            'username': row[1],
            'first_name': row[2],
            'last_name': row[3],
            'email': row[4],
            'phone': row[5],
            'address': row[6]
        })
    
    return render_template('customers.html', 
                         customers=customers, 
                         username=session['username'],
                         page_title='Customers Management',
                         active_page='customers')

def export_merk_csv(mysql):
    if 'username' not in session:
        return redirect(url_for('login'))
    
    if session.get('admin') != 1:
        flash("You do not have admin privileges")
        return redirect(url_for('home'))
    
    try:

        cur = mysql.connection.cursor()
        cur.execute("SELECT merk_id, nama_merk, jumlahpenjualan, keuntungan FROM merk ORDER BY merk_id")
        merks_data = cur.fetchall()
        cur.close()

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Merk Data"
        
        # Add headers
        headers = ['ID', 'Nama Merk', 'Jumlah Penjualan', 'Keuntungan (Million)']
        ws.append(headers)
        
        # Style header row
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        # Add data rows
        for row in merks_data:
            ws.append([row[0], row[1], row[2], row[3]])
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'merk_export_{timestamp}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        flash(f"Error exporting data: {str(e)}", "error")
        return redirect(url_for('merk_list'))

def order_list(mysql):
    if 'username' not in session:
        return redirect(url_for('login'))
    
    if session.get('admin') != 1:
        flash("You do not have admin privileges")
        return redirect(url_for('home'))
    
    action = request.args.get('action', None)
    order_id = request.args.get('order_id', None)
    order = None
    show_details = False
    
    # Handle status update
    if request.method == 'POST':
        order_id_form = request.form.get('order_id')
        new_status = request.form.get('status')
        
        if order_id_form and new_status:
            try:
                cur = mysql.connection.cursor()
                cur.execute("UPDATE history SET status_produk = %s WHERE history_id = %s", (new_status, order_id_form))
                mysql.connection.commit()
                cur.close()
                flash("Order status updated successfully", "success")
            except Exception as e:
                flash(f"Error updating order: {str(e)}", "error")
            
            return redirect(url_for('order_list'))
    
    if action == 'detail' and order_id:
        show_details = True
        cur = mysql.connection.cursor()
        cur.execute("""
            SELECT h.history_id, h.user_id, u.username, p.nama_produk, h.banyak, 
                   h.jumlah_pembayaran, h.tanggal_pembelian, h.status_produk
            FROM history h
            JOIN users u ON h.user_id = u.id
            JOIN produk p ON h.produk_id = p.produk_id
            WHERE h.history_id = %s
        """, (order_id,))
        row = cur.fetchone()
        cur.close()
        
        if row:
            order = {
                'history_id': row[0],
                'user_id': row[1],
                'username': row[2],
                'nama_produk': row[3],
                'banyak': row[4],
                'jumlah_pembayaran': row[5],
                'tanggal_pembelian': row[6],
                'status_produk': row[7]
            }
        else:
            flash("Order not found", "error")
    
    # Get all orders
    cur = mysql.connection.cursor()
    cur.execute("""
        SELECT h.history_id, u.username, p.nama_produk, h.banyak, 
               h.jumlah_pembayaran, h.tanggal_pembelian, h.status_produk
        FROM history h
        JOIN users u ON h.user_id = u.id
        JOIN produk p ON h.produk_id = p.produk_id
        ORDER BY h.tanggal_pembelian DESC
    """)
    orders_data = cur.fetchall()
    cur.close()
    
    orders = []
    for row in orders_data:
        orders.append({
            'history_id': row[0],
            'username': row[1],
            'nama_produk': row[2],
            'banyak': row[3],
            'jumlah_pembayaran': row[4],
            'tanggal_pembelian': row[5],
            'status_produk': row[6]
        })
    
    return render_template('order.html',
                         orders=orders,
                         order=order,
                         show_details=show_details,
                         username=session['username'],
                         page_title='Orders Management',
                         active_page='orders')

def product_list(mysql):
    if 'username' not in session:
        return redirect(url_for('login'))
    
    if session.get('admin') != 1:
        flash("You do not have admin privileges")
        return redirect(url_for('home'))
    
    action = request.args.get('action', None)
    product_id = request.args.get('product_id', None)
    product = None
    merks = []
    show_form = False
    
    # Get all merks for dropdown
    cur = mysql.connection.cursor()
    cur.execute("SELECT merk_id, nama_merk FROM merk ORDER BY nama_merk")
    merks_data = cur.fetchall()
    cur.close()
    
    for row in merks_data:
        merks.append({
            'merk_id': row[0],
            'nama_merk': row[1]
        })
    
    if action == 'create':
        show_form = True
    elif action == 'edit' and product_id:
        show_form = True
        cur = mysql.connection.cursor()
        cur.execute("SELECT produk_id, nama_produk, harga, stok, merk_id FROM produk WHERE produk_id = %s", (product_id,))
        row = cur.fetchone()
        cur.close()
        
        if row:
            product = {
                'produk_id': row[0],
                'nama_produk': row[1],
                'harga': row[2],
                'stok': row[3],
                'merk_id': row[4]
            }
        else:
            flash("Product not found", "error")
    
    if request.method == 'POST':
        nama_produk = request.form.get('nama_produk', '').strip()
        harga = request.form.get('harga', 0)
        stok = request.form.get('stok', 0)
        merk_id = request.form.get('merk_id', None)
        
        if not nama_produk:
            flash("Product name is required", "error")
            return redirect(url_for('product_list', action=action, product_id=product_id))
        
        try:
            cur = mysql.connection.cursor()
            if product_id:
                cur.execute("UPDATE produk SET nama_produk = %s, harga = %s, stok = %s, merk_id = %s WHERE produk_id = %s",
                            (nama_produk, harga, stok, merk_id, product_id))
                flash("Product updated successfully", "success")
            else:
                cur.execute("INSERT INTO produk (nama_produk, harga, stok, merk_id) VALUES (%s, %s, %s, %s)",
                            (nama_produk, harga, stok, merk_id))
                flash("Product created successfully", "success")
            
            mysql.connection.commit()
            cur.close()
            
            return redirect(url_for('product_list'))
        except Exception as e:
            flash(f"Error: {str(e)}", "error")
            return redirect(url_for('product_list', action=action, product_id=product_id))
    
    # Get all products
    cur = mysql.connection.cursor()
    cur.execute("""
        SELECT p.produk_id, p.nama_produk, p.harga, p.stok, p.merk_id, m.nama_merk
        FROM produk p
        LEFT JOIN merk m ON p.merk_id = m.merk_id
        ORDER BY p.produk_id DESC
    """)
    products_data = cur.fetchall()
    cur.close()
    
    products = []
    for row in products_data:
        products.append({
            'produk_id': row[0],
            'nama_produk': row[1],
            'harga': row[2],
            'stok': row[3],
            'merk_id': row[4],
            'nama_merk': row[5] if row[5] else 'N/A'
        })
    
    return render_template('product.html',
                         products=products,
                         product=product,
                         merks=merks,
                         show_form=show_form,
                         username=session['username'],
                         page_title='Products Management',
                         active_page='products')

def product_delete(mysql, product_id):
    if 'username' not in session:
        return redirect(url_for('login'))
    
    if session.get('admin') != 1:
        flash("You do not have admin privileges")
        return redirect(url_for('home'))
    
    try:
        cur = mysql.connection.cursor()
        cur.execute("DELETE FROM produk WHERE produk_id = %s", (product_id,))
        mysql.connection.commit()
        cur.close()
        
        flash("Product deleted successfully", "success")
    except Exception as e:
        flash(f"Error: {str(e)}", "error")
    
    return redirect(url_for('product_list'))